import os
import markdownify
from docx import Document
import pdfplumber
import re
import csv
from openpyxl import load_workbook
import chardet
from typing import List, Tuple, Any, cast
import pytesseract as tess

def clean_text(text: str) -> str:
    text = re.sub(r'([a-z])([A-Z])', r'\1 \2', text)
    text = re.sub(r'([a-zA-Z])(\d)', r'\1 \2', text)
    text = re.sub(r'(\d)([a-zA-Z])', r'\1 \2', text)
    text = re.sub(r'([.,!?])([a-zA-Z])', r'\1 \2', text)
    return text

def _table_to_markdown(header: List[Any], rows: List[List[Any]]) -> str:
    if not header and not rows:
        return ""
    
    markdown_table = f"| {' | '.join(map(str, header))} |\n"
    markdown_table += f"|{'|'.join(['---'] * len(header))}|\n"
    
    for row in rows:
        markdown_table += f"| {' | '.join(map(str, row))} |\n"
        
    return markdown_table

def convert_to_markdown(input_file: str, output_file: str) -> Tuple[bool, str]:
    file_extension: str = os.path.splitext(input_file)[1].lower()
    content: str = ""
    markdown_content: str = ""
    
    try:
        if file_extension in ['.txt', '.csv']:
            with open(input_file, 'rb') as f:
                raw_data = f.read()
                detected = chardet.detect(raw_data)
                encoding = detected['encoding'] if detected['confidence'] > 0.5 else 'utf-8'
                print(f"   - Detected encoding for {os.path.basename(input_file)}: {encoding}", flush=True)

            with open(input_file, 'r', encoding=encoding, errors='replace') as f:
                if file_extension == '.txt':
                    content = f.read()
                elif file_extension == '.csv':
                    reader = csv.reader(f)
                    header = next(reader)
                    rows = [row for row in reader]
                    markdown_content = _table_to_markdown(header, rows)
        
        elif file_extension == '.docx':
            doc = Document(input_file)
            content = "\n\n".join([para.text for para in doc.paragraphs if para.text.strip()])
        
        elif file_extension == '.pdf':
            with pdfplumber.open(input_file) as pdf:
                for page in pdf.pages:
                    text: str = page.extract_text(layout=True, x_density=300, y_density=300)
                    if text and text.strip():
                        content += text + "\n\n"
                    else:
                        try:
                            img = page.to_image(resolution=300)
                            text = cast(str, tess.image_to_string(img.original))  # type: ignore
                            content += text + "\n\n"
                        except Exception as e:
                            content += f"[OCR failed on page: {str(e)}]\n\n"
        
        elif file_extension == '.xlsx':
            workbook = load_workbook(filename=input_file)
            sheet = workbook.active
            if sheet is None:
                raise ValueError("No active sheet in workbook")
            
            header = [cell.value for cell in sheet[1]]
            rows: List[List[Any]] = []
            for row in sheet.iter_rows(min_row=2):
                rows.append([cell.value for cell in row])
            markdown_content = _table_to_markdown(header, rows)

        elif file_extension in ['.html', '.js', '.css', '.py', '.cpp', '.c', '.java', '.cs', '.ts', '.json', '.xml', '.log', '.sql', '.php', '.rb', '.go', '.rs', '.yml', '.yaml', '.ini', '.cfg', '.conf', '.sh', '.bat', '.ps1', '.lua', '.pl', '.tcl', '.r', '.m', '.swift', '.kt', '.scala', '.dart', '.hs', '.ml', '.fs', '.vb', '.asm', '.s', '.tex', '.bib', '.sty']:
            with open(input_file, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
            markdown_content = f"```{file_extension[1:]}\n{content}\n```"
        
        else:
            raise ValueError(f"Unsupported file type for conversion: {file_extension}")
        
        if not markdown_content:
            markdown_content = markdownify.markdownify(content, heading_style="ATX")

        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(markdown_content)
        
        return True, f"Successfully converted {input_file} to {output_file}"
    
    except Exception as e:
        return False, f"Error converting {input_file}: {str(e)}"